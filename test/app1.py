import os  
from flask import Flask,render_template, request, jsonify, redirect, url_for, session ,send_from_directory
import mysql.connector
from mysql.connector import pooling
from flask_session import Session
import pymysql.cursors 
import cv2
import numpy as np
import io
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, Border, Side
from flask import send_file
import json
from datetime import datetime

app = Flask(__name__)  
app.secret_key = 'your_secret_key'  

UPLOAD_FOLDER = os.path.join(os.path.expanduser("~"), "Desktop", "AssessmentPictures")
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

UPLOAD_IMG = os.path.join(os.path.expanduser("~"), "Desktop", "RemedyPictures")
app.config["UPLOAD_IMG"] = UPLOAD_IMG

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

if not os.path.exists(UPLOAD_IMG):
    os.makedirs(UPLOAD_IMG)

# Connect to MySQL
#def get_db_connection():
#    return mysql.connector.connect(
#        host="localhost",  # Connecting to localhost
#        user="root",
#        password="nare@2058",
#        database="remedydb",
#        autocommit=True,  # Prevents timeout issues
#        connection_timeout=300000  # Keeps connection alive
# )

db_pool = pooling.MySQLConnectionPool(
    pool_name="mypool",
    pool_size=10,  # Adjust based on workload
    host="localhost",  # Or server IP if hosted remotely
    user="root",
    password="nare@2058", 
    database="remedydb",
    autocommit=True,  # Prevents timeout issues
    connection_timeout=60 # 60 seconds is ideal
   
)
def get_db_connection():
    return db_pool.get_connection()


@app.route('/export_excel')
def export_excel():
    # Your real data fetching logic goes here
    assessment_data = [
        {
            "Sl No": 1,
            "Table ID": "C31S19",
            "Pile No": "Pile1",
            "Date of Assessment": "2025-04-14",
            "CASE 1": ["Yes", "No", "Yes"],
            "CASE 2": ["No", "Yes"],
            "CASE 3": ["Yes", "No", "Yes"],
            "CASE 4": ["Yes"],
            "Assessment Status": "Completed",
            "Classification": "Moderate",
            "Recommendation": "Paint Coating",
            "Remarks": "Rust mostly on base",
            "Picture Path": "static/images/AS00037_C31S19_Pile1_side1.jpg"
        }
    ]

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assessment Report"

    header1 = ["Sl. No.", "Table ID", "Pile No", "Date of Assessment", "Criteria/Parameters", "", "", "", "", "", "", "", "", "", "Assessment Status", "Classification as per Indisolar", "Recommendation as per OE", "Remarks", "Picture"]
    header2 = ["", "", "", "", "CASE 1", "", "", "CASE 2", "", "CASE 3", "", "", "CASE 4", "", "", "", "", ""]
    header3 = ["", "", "", "", "The pile completely lost galvanized coating", "Damages in Bare Steel", "Several rusty areas on metal surface", "The pile Partially or completely lost galvanized coating", "Bare Steel beneath exposed with several red rusty spots or rusty areas", "Has the pile partially lost galvanized coating", "Bare steel unaffected", "Few red spots or rusty areas on metal surface", "Zinc Carbonates on surface with no red spots or rusty areas (Visual inspection)", "", "", "", "", ""]

    ws.append(header1)
    ws.append(header2)
    ws.append(header3)

    def merge_cells(ws):
        ws.merge_cells("A1:A3")
        ws.merge_cells("B1:B3")
        ws.merge_cells("C1:C3")
        ws.merge_cells("D1:D3")
        ws.merge_cells("E1:N1")
        ws.merge_cells("E2:G2")
        ws.merge_cells("H2:I2")
        ws.merge_cells("J2:L2")
        ws.merge_cells("M2:M3")
        ws.merge_cells("N2:N3")
        ws.merge_cells("O1:O3")
        ws.merge_cells("P1:P3")
        ws.merge_cells("Q1:Q3")
        ws.merge_cells("R1:R3")

    def format_cells():
        for row in ws.iter_rows(min_row=1, max_row=3):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(bold=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

    merge_cells(ws)
    format_cells()

    row_num = 4
    for entry in assessment_data:
        row = [
            entry["Sl No"],
            entry["Table ID"],
            entry["Pile No"],
            entry["Date of Assessment"],
            *entry["CASE 1"],
            *entry["CASE 2"],
            *entry["CASE 3"],
            *entry["CASE 4"],
            entry["Assessment Status"],
            entry["Classification"],
            entry["Recommendation"],
            entry["Remarks"]
        ]
        ws.append(row)

        # Add image
        img_path = entry["Picture Path"]
        if os.path.exists(img_path):
            img = ExcelImage(img_path)
            img.width = 80
            img.height = 60
            ws.add_image(img, f"R{row_num}")
        else:
            ws[f"R{row_num}"] = "Image not found"

        row_num += 1

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        from openpyxl.utils import get_column_letter

        for col_idx, col in enumerate(ws.columns, start=1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2


    # Save to in-memory buffer
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(file_stream, as_attachment=True, download_name="Assessment_Report.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
Session(app)

from flask import send_from_directory

# Folder where you want to save the file (renamed as requested)
assessment_html_changes = r'C:\Users\LENOVO\Desktop\HTMLreport'
@app.route("/save_assessment_changes", methods=["POST"])
def save_assessment_changes():
    try:
        data = request.json.get("data", [])

        # Create a timestamped filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"assessment_changes_{timestamp}.json"

        # Full path
        file_path = os.path.join(assessment_html_changes, filename)

        # Ensure folder exists
        os.makedirs(assessment_html_changes, exist_ok=True)

        # Save to file
        with open(file_path, "w") as f:
            json.dump(data, f, indent=4)

        return jsonify({"message": "Assessment changes saved successfully!", "file": filename})
    
    except Exception as e:
        print("Error:", e)
        return jsonify({"message": "Error saving data."}), 500

# Set path where your images are saved
ASSESSMENT_PIC_FOLDER = r'C:\Users\LENOVO\Desktop\AssessmentPictures'

@app.route('/AssessmentPictures/<folder>/<filename>')
def serve_assessment_image(folder, filename):
    return send_from_directory(os.path.join(ASSESSMENT_PIC_FOLDER, folder), filename)

# Set path where your remedy images are saved
REMEDY_PIC_FOLDER = r'C:\Users\LENOVO\Desktop\RemedyPictures'
@app.route('/RemedyPictures/<folder>/<filename>')
def serve_remedy_image(folder, filename):
    return send_from_directory(os.path.join(REMEDY_PIC_FOLDER, folder), filename)
    
# Define the base folder path for images and case folders
BASE_FOLDER = 'static/images'  # Base folder path for images
CASE_FOLDERS = ['case1', 'case2', 'case3', 'case4']
def mask_metal_region(image):
    """Mask the background and keep only the metal pile region using color thresholding."""
    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    # Metal piles are typically grayish, so target low saturation
    lower_metal = np.array([0, 0, 60])
    upper_metal = np.array([180, 70, 255])

    # Create mask where metal is likely present
    metal_mask = cv2.inRange(hsv, lower_metal, upper_metal)

    # Apply mask to keep only metal region
    result = cv2.bitwise_and(image, image, mask=metal_mask)

    return result

def calculate_rust_percentage(base_img, test_img):
    # Convert to HSV
    hsv_base = cv2.cvtColor(base_img, cv2.COLOR_BGR2HSV)
    hsv_test = cv2.cvtColor(test_img, cv2.COLOR_BGR2HSV)

    # Define a range for rust color in HSV (you can adjust as needed)
    lower_red = np.array([0, 50, 50])
    upper_red = np.array([10, 255, 255])

    # Create masks for rust areas
    base_mask = cv2.inRange(hsv_base, lower_red, upper_red)
    test_mask = cv2.inRange(hsv_test, lower_red, upper_red)

    # Find overlapping rust regions (base vs test)
    overlap = cv2.bitwise_and(base_mask, test_mask)
    match_pixels = np.count_nonzero(overlap)
    total_pixels = np.count_nonzero(base_mask)

    if total_pixels == 0:
        return 0.0  # Avoid division by zero if no rust in base

    match_percent = (match_pixels / total_pixels) * 100
    return match_percent
def detect_rust_and_damage_percentage(image):
    """Detect rust (red color) and damage (bare iron) percentage in the image."""
    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    # Define red ranges in HSV for rust detection
    lower_red1 = np.array([0, 70, 50])
    upper_red1 = np.array([10, 255, 255])
    lower_red2 = np.array([160, 70, 50])
    upper_red2 = np.array([180, 255, 255])

    # Define low saturation and high value for detecting bare iron (damaged surface)
    lower_bare = np.array([0, 0, 100])  # Low saturation, high value (bare metal)
    upper_bare = np.array([180, 50, 255])

    # Mask for rust (red) and bare metal (damage)
    mask_rust1 = cv2.inRange(hsv, lower_red1, upper_red1)
    mask_rust2 = cv2.inRange(hsv, lower_red2, upper_red2)
    mask_bare = cv2.inRange(hsv, lower_bare, upper_bare)

    # Combine the masks for rust and bare metal (damage)
    rust_pixels = cv2.countNonZero(mask_rust1) + cv2.countNonZero(mask_rust2)
    damage_pixels = cv2.countNonZero(mask_bare)
    total_pixels = image.shape[0] * image.shape[1]

    rust_percentage = (rust_pixels / total_pixels) * 100
    damage_percentage = (damage_pixels / total_pixels) * 100

    return rust_percentage, damage_percentage

def get_max_values_for_case(case_folder, base_image):
    """Get the max rust and damage percentage for a given case folder."""
    max_rust = 0
    max_damage = 0
    
    # Go through each side image in the case folder (side1.jpg, side2.jpg, side3.jpg, side4.jpg)
    for side in range(1, 5):
        side_image_path = os.path.join(BASE_FOLDER, case_folder, f"side{side}.jpg")
        
        # Check if the side image exists
        if not os.path.exists(side_image_path):
            continue
        
        # Read the side image
        side_img = cv2.imread(side_image_path)
        
        # Calculate rust and damage percentages
        rust_percent, damage_percent = detect_rust_and_damage_percentage(side_img)
        
        # Update the max values if current values are higher
        max_rust = max(max_rust, rust_percent)
        max_damage = max(max_damage, damage_percent)
    
    return max_rust, max_damage

@app.route('/analyze_corrosion', methods=['POST'])
def analyze_corrosion():
    uploaded_images = []
    rust_scores_by_case = {case: [] for case in CASE_FOLDERS}

    # Step 1: Read uploaded images (side1 to side4)
    for i in range(1, 5):
        file = request.files.get(f'side{i}')
        if not file:
            print(f"❌ Missing uploaded image: side{i}")
            return jsonify({"error": f"Missing image: side{i}"}), 400

        # print(f"✅ Received: {file.filename}")
        file_bytes = np.frombuffer(file.read(), np.uint8)
        img = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
        if img is None:
            return jsonify({"error": f"Cannot decode image: side{i}"}), 400
        img = mask_metal_region(img)
        img = cv2.resize(img, (300, 300))


        uploaded_images.append(img)

    # Step 2: Compare each uploaded image to corresponding base image in each case
    for case in CASE_FOLDERS:
        # print(f"📁 Analyzing against: {case}")
        case_scores = []
        for i in range(1, 5):  # side1 to side4
            base_img_path = os.path.join(BASE_FOLDER, case, f'side{i}.jpg')
            if not os.path.exists(base_img_path):
                # print(f"⚠️ Missing base image: {base_img_path}")
                case_scores.append(0)
                continue

            base_img = cv2.imread(base_img_path)
            base_img = cv2.resize(base_img, (300, 300))

            # Calculate rust/damage % between base and uploaded
            rust_percent = calculate_rust_percentage(base_img, uploaded_images[i - 1])
            # print(f"🧪 Rust match {case} side{i}: {rust_percent:.2f}%")
            case_scores.append(round(rust_percent))

        rust_scores_by_case[case] = case_scores

    # Step 3: Return average rust match per case + individual side details
    rust_summary = [round(np.mean(rust_scores_by_case[case])) for case in CASE_FOLDERS]
    # print("📊 Final rust match summary:", rust_summary)

    return jsonify({
        "damage_per_image": rust_summary,
        "details": {
            case: {
                f"side{i+1}": rust_scores_by_case[case][i]
                for i in range(4)
            } for case in CASE_FOLDERS
        }
    })

@app.route('/get_tasks', methods=['GET'])
def get_tasks():
    user_id = request.args.get('user_id')
    task_date = request.args.get('task_date')

    if not user_id or not task_date:
        return jsonify({"success": False, "message": "Missing parameters"}), 400

    # ✅ Use DictCursor to get row as dictionary
    connection = get_db_connection()
    cursor = connection.cursor(pymysql.cursors.DictCursor)

    cursor.execute("""
        SELECT * FROM assessment
        WHERE `User ID` = %s AND `Task Date` = %s
    """, (user_id, task_date))

    rows = cursor.fetchall()
    connection.close()

    return jsonify(rows)

    
# ✅ Set your custom image storage path
UPLOAD_FOLDER = r"C:\Users\LENOVO\Desktop\AssessmentPictures"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route("/upload_single_image", methods=["POST"])
def upload_single_image():
    image = request.files.get("image")
    folder = request.form.get("folder")  # Assessment ID (used as subfolder)
    filename = request.form.get("filename")

    if not image or not folder or not filename:
        return "❌ Missing image or metadata", 400

    # ✅ Create subfolder based on Assessment ID
    folder_path = os.path.join(UPLOAD_FOLDER, folder)
    os.makedirs(folder_path, exist_ok=True)

    # ✅ Save the image
    image_path = os.path.join(folder_path, filename)
    image.save(image_path)

    return f"✅ Image saved to {folder}/{filename}", 200
    
@app.route('/upload_tasks_to_pc', methods=['POST'])
def upload_tasks_to_pc():
    data = request.json
    uploaded_tasks = data.get("tasks", [])

    connection = get_db_connection()  # Your MySQL connection function
    cursor = connection.cursor(dictionary=True)  # ✅ mysql.connector-compatible

    updated = 0
    inserted = 0
    skipped = 0

    for row in uploaded_tasks:
        assessment_id = row[0]

        # ✅ Update Picture Location path for PC
        picture_location = rf"C:\Users\LENOVO\Desktop\AssessmentPictures\{assessment_id}"
        row[15] = picture_location  # Overwrite with correct PC path

        # Check if record exists
        cursor.execute("SELECT * FROM assessment WHERE `Assessment ID` = %s", (assessment_id,))
        existing = cursor.fetchone()

        if existing:
            # Compare each relevant field before deciding to update
            fields_to_compare = {
                "Task Date": row[5],
                "Allotted Date": row[6],
                "Allotted By": row[7],
                "Date Completed": row[8],
                "Assessment Status": row[9],
                "Assessment case": row[10],
                "Picture1 Name": row[11],
                "Picture2 Name": row[12],
                "Picture3 Name": row[13],
                "Picture4 Name": row[14],
                "Picture Location": row[15]
            }

            changes_required = False
            for key, new_value in fields_to_compare.items():
                if str(existing[key]) != str(new_value):
                    changes_required = True
                    break

            if changes_required:
                cursor.execute("""
                    UPDATE assessment SET
                        `Task Date` = %s,
                        `Allotted Date` = %s,
                        `Allotted By` = %s,
                        `Date Completed` = %s,
                        `Assessment Status` = %s,
                        `Assessment case` = %s,
                        `Picture1 Name` = %s,
                        `Picture2 Name` = %s,
                        `Picture3 Name` = %s,
                        `Picture4 Name` = %s,
                        `Picture Location` = %s
                    WHERE `Assessment ID` = %s
                """, (
                    row[5], row[6], row[7], row[8], row[9], row[10],
                    row[11], row[12], row[13], row[14], row[15],
                    assessment_id
                ))
                updated += 1
            else:
                skipped += 1  # No update needed
        else:
            # Insert new record
            cursor.execute("""
                INSERT INTO assessment (
                    `Assessment ID`, `Area ID`, `User ID`, `Table ID`, `Pile No`, `Task Date`, 
                    `Allotted Date`, `Allotted By`, `Date Completed`, `Assessment Status`, 
                    `Assessment case`, `Picture1 Name`, `Picture2 Name`, `Picture3 Name`, 
                    `Picture4 Name`, `Picture Location`
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, row)
            inserted += 1

    connection.commit()
    cursor.close()
    connection.close()

    return jsonify({
        "success": True,
        "message":f"Data Uploaded successfully --\n{inserted} inserted, {updated} updated, {skipped} skipped."
    })

# on your PC Flask server (app.py)
@app.route("/discover")
def discover():
    return jsonify({"status": "available", "ip": request.host.split(':')[0]})

@app.route('/sendtdpc')
def send_today_pc():
    return render_template('sendtdpc.html')

@app.route('/service-worker.js')
def serve_worker():
    return send_from_directory('static/js', 'service-worker.js')
    
@app.route('/')
def index():
    return render_template('login.html')

@app.route('/A001hotspot')
def A001hotspot():
    return render_template('A001hotspot.html')

@app.route('/A002hotspot')
def A002hotspot():
    return render_template('A002hotspot.html')

@app.route('/A003hotspot')
def A003hotspot():
    return render_template('A003hotspot.html')

@app.route('/A004hotspot')
def A004hotspot():
    return render_template('A004hotspot.html')

@app.route('/A005hotspot')
def A005hotspot():
    return render_template('A005hotspot.html')

@app.route('/A001remedyhotspot')
def A001_remedy_hotspot():
    return render_template('A001remedyhotspot.html')

@app.route('/A002remedyhotspot')
def A002_remedy_hotspot():
    return render_template('A002remedyhotspot.html')

@app.route('/A003remedyhotspot')
def A003_remedy_hotspot():
    return render_template('A003remedyhotspot.html')

@app.route('/A004remedyhotspot')
def A004_remedy_hotspot():
    return render_template('A004remedyhotspot.html')

@app.route('/A005remedyhotspot')
def A005_remedy_hotspot():
    return render_template('A005remedyhotspot.html')

# Route for the home page
@app.route('/dashboard', methods=['GET', 'POST'])
def dashboard():
    if 'email' not in session:
        return redirect('/login')  # Redirect to login if not authenticated

    user_type = session.get('user_type', 'Normal User')
    username = session.get('username')  # Get the username from the session
    return render_template('dashboard.html', user_type=user_type, username=username)

@app.context_processor
def inject_user():
    full_name = session.get('username', '')  
    first_name = full_name.split()[0] if full_name else '' 
    return {
        'username': first_name,  
        'user_type': session.get('user_type', 'Normal User')
    }

# for create route 
@app.route('/site')
def site():
    return render_template('site.html')

# route to select the pile from the image map
@app.route('/image-map')
def image_map():
    return render_template('area1hotspot.html')

@app.route('/customer')
def customer():
    return render_template('customer.html')

@app.route('/userform')
def userform():
    return render_template('userform.html')

@app.route('/area')
def area():
    return render_template('area.html')

@app.route('/rows')
def rows():
    return render_template('rows.html')

@app.route('/tables')
def tables():
    return render_template('tables.html')

@app.route('/piles')
def piles():
    return render_template('pile.html')

@app.route('/assessment')
def assessment():
    return render_template('assessment.html')

@app.route('/remedy')
def remedy():
    return render_template('remedy.html')

@app.route('/inventory')
def inventory():
    return render_template('inventory.html')

@app.route('/invtrans')
def invtrans():
    return render_template('invtrans.html')

#@app.route('/quality')
#def quality():
    #return render_template('quality.html')

@app.route('/reports')
def reports():
    return render_template('assreports.html')

@app.route('/reporthtml')
def reporthtml():
    return render_template('reporthtml.html')
    
@app.route('/remedyreports')
def remedyreports():
    return render_template('remedyreports.html')

@app.route('/profile', methods=['GET'])
def profile():
    if 'email' not in session:  
        return redirect('/login')  

    email = session['email']  # Retrieve the email from the session
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

   
    cursor.execute("SELECT * FROM users WHERE `Email` = %s", (email,))
    user = cursor.fetchone()

    if not user:  # Handle the case where no user is found
        return redirect('/login')  # Redirect to login if the user is not found

    # Pass the user data to the profile template
    return render_template('profile.html', user=user)


@app.route('/user_log')
def user_log():
    return render_template('user_log.html')

@app.route('/comments')
def comments():
    return render_template('comments.html')

@app.route('/area1remedyhotspot')
def area12hotspot():
    return render_template('area1remedyhotspot.html')

################################################################

#for update route

@app.route('/updateusers')
def update_users():
   
    return render_template('updateusers.html')

@app.route('/updatesite')
def update_site():
    
    return render_template('updatesite.html')

@app.route('/updatecustomer')
def update_customer():
   
    return render_template('updatecustomer.html')

@app.route('/updateinventory')
def update_inventory():
    
    return render_template('updateinventory.html')

@app.route('/updateinvtrans')
def update_invtrans():
   
    return render_template('updateinvtrans.html')

@app.route('/updatearea')
def update_area():
    return render_template('updatearea.html')

@app.route('/updatetable')
def update_table():
    return render_template('updatetable.html')

@app.route('/updatepile')
def update_pile():
    return render_template('updatepile.html')

@app.route('/updaterow')
def update_row():
    return render_template('updaterow.html')

@app.route('/updateassmnt')
def update_assmnt():
    return render_template('updateassmnt.html')

@app.route('/updateremedy')
def update_remedy():
    return render_template('updateremedy.html')

###########################################################
@app.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    email = data.get('email')
    password = data.get('password')

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("SELECT * FROM users WHERE email = %s AND password = %s", (email, password))
        user = cursor.fetchone()

        if user:
            # Store necessary details in the session
            session['email'] = user['Email']  
            session['username'] = user['User Name']  
            session['user_type'] = user['User Type']

            return jsonify({"success": True, "message": "Login successful"})
        else:
            return jsonify({"success": False, "message": "Invalid credentials"})

    except Exception as e:
        return jsonify({"success": False, "message": "An error occurred", "error": str(e)})

    finally:
        cursor.close()
        connection.close()
 

# Route for user creation
@app.route('/create_user', methods=['POST'])
def create_user():
    data = request.get_json()
    name = data.get('name')
    email = data.get('email')
    password = data.get('password')

    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500  # Handle connection failure

    cursor = connection.cursor()
    try:
        query = "INSERT INTO loginusers (name, email, password) VALUES (%s, %s, %s)"
        cursor.execute(query, (name, email, password))
        connection.commit()
        return jsonify({"success": True, "message": "User created successfully"})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error creating user: {e}"}), 500
    finally:
        cursor.close()  # Ensure cursor is closed
        connection.close()  # Ensure connection is closed

# Route for logging out
@app.route('/logout', methods=['POST'])
def logout():
    session.pop('username', None)  # Remove the username from the session
    return redirect(url_for('index'))

@app.route('/submit_siteform', methods=['POST'])
def submit_siteform():
    site_name = request.form.get('site_name')
    site_location = request.form.get('location')
    site_owner = request.form.get('site_owner_name')
    site_gps = request.form.get('site_gps')

    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500  # Handle DB connection failure

    cursor = connection.cursor()
    try:
        # Fetch the current maximum Site ID
        cursor.execute("SELECT `Site ID` FROM `Site` ORDER BY `Site ID` DESC LIMIT 1")
        result = cursor.fetchone()
        
        next_number = int(result[0][1:]) + 1 if result else 1
        new_site_id = f"S{next_number:03d}"  

        query = """
        INSERT INTO `Site` (`Site ID`, `Cust ID`, `Site Name`, `Site Location`, `Site Owner Name`, `Site GPS`)
        VALUES (%s, %s, %s, %s, %s, %s)
        """
        cursor.execute(query, (new_site_id, "", site_name, site_location, site_owner, site_gps))
        connection.commit()

        return jsonify({"success": True, "message": "Site information saved successfully"})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error saving site information: {e}"}), 500
    finally:
        cursor.close()  
        connection.close()  

@app.route('/submit_customerform', methods=['POST'])
def submit_customerform():
    name = request.form.get('name')
    address = request.form.get('address')
    contact_person = request.form.get('contact_person')
    website = request.form.get('website')
    phone_no = request.form.get('phone_no')
    country = request.form.get('country')

    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    cursor = connection.cursor()
    try:
        # Fetch the current maximum Cust ID
        cursor.execute("SELECT `Cust ID` FROM `Customer` ORDER BY `Cust ID` DESC LIMIT 1")
        result = cursor.fetchone()
        
        next_number = int(result[0][1:]) + 1 if result else 1
        new_cust_id = f"C{next_number:03d}"  

        query = """
        INSERT INTO Customer 
        (`Cust ID`, `Customer Name`, `Customer Address`, `Contact Person`, `Customer Website`, `Phone No`, `Country`)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(query, (new_cust_id, name, address, contact_person, website, phone_no, country))
        connection.commit()

        return jsonify({"success": True, "message": "Customer information saved successfully"})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error saving customer information: {e}"}), 500
    finally:
        cursor.close()  
        connection.close()  

@app.route('/submit_user_form', methods=['POST'])
def submit_userform():
    # Retrieve form data
    user_name = request.form.get('user_name')
    user_type = request.form.get('user_type')
    designation = request.form.get('designation')
    phone_no = request.form.get('phone_no')
    reports_to = request.form.get('reports_to')
    date_created = request.form.get('date_created')
    site_id = request.form.get('site_id')  
    email = request.form.get('gmail_address') 
    password = request.form.get('create_password')
    confirm_password = request.form.get('confirm_password') 

    if password != confirm_password:
        return jsonify({"success": False, "message": "Passwords do not match!"})

    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500  # Handle DB connection failure

    cursor = connection.cursor()
    try:
        # Check if the Site ID exists
        cursor.execute("SELECT `Site ID` FROM `site` WHERE `Site ID` = %s", (site_id,))
        if not cursor.fetchone():
            return jsonify({"success": False, "message": "Invalid Site ID selected."})

        # Check if Email already exists
        cursor.execute("SELECT `Email` FROM `users` WHERE `Email` = %s", (email,))
        if cursor.fetchone():
            return jsonify({"success": False, "message": "Email already exists. Please use a different email."})

        # Fetch last User ID
        cursor.execute("SELECT `User ID` FROM `users` ORDER BY `User ID` DESC LIMIT 1")
        result = cursor.fetchone()
        next_number = int(result[0][1:]) + 1 if result else 1
        new_user_id = f"U{next_number:03d}"  

        query = """
        INSERT INTO `users` (`User ID`, `Site ID`, `User Name`, `User Type`, `User Designation`, `User Phone number`, `Reports To`, `Date Created`, `Date Removed`, `Email`, `Password`)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(query, (new_user_id, site_id, user_name, user_type, designation, phone_no, reports_to, date_created, None, email, password))
        connection.commit()

        return jsonify({"success": True, "message": "User information saved successfully"})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error saving user information: {e}"}), 500
    finally:
        cursor.close()
        connection.close()

@app.route('/submit_area_form', methods=['POST'])
def submit_area_form():
    # Retrieve form data
    location = request.form.get('location')
    gps = request.form.get('gps')
    if not location or not gps:
        return jsonify({"success": False, "message": "All fields are required."})

    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    cursor = connection.cursor()

    try:
        # Fetch the current maximum Area ID
        cursor.execute("SELECT `Area ID` FROM `areas` ORDER BY `Area ID` DESC LIMIT 1")
        result = cursor.fetchone()

        # Determine the next Area ID
        next_number = int(result[0][1:]) + 1 if result else 1
        new_area_id = f"A{next_number:03d}"

        query = """
        INSERT INTO `areas` (`Area ID`, `Location`, `GPS`)
        VALUES (%s, %s, %s)
        """
        cursor.execute(query, (new_area_id, location, gps)) 
        connection.commit()

        return jsonify({"success": True, "message": "Area information saved successfully"})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error saving area information: {e}"}), 500
    finally:
        cursor.close()
        connection.close()

@app.route('/submit_user_log_form', methods=['POST'])
def submit_user_log_form():
    user_id = request.form.get('user_id')
    date_logged_in = request.form.get('date_logged_in')
    date_logged_out = request.form.get('date_logged_out')

    if not user_id or not date_logged_in:
        return jsonify({"success": False, "message": "User ID and Date Logged In are required."}), 400

    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    cursor = connection.cursor()

    try:
        query = """
        INSERT INTO `user_log` (`User ID`, `Date Logged in`, `Date Logged out`)
        VALUES (%s, %s, %s)
        """
        cursor.execute(query, (user_id, date_logged_in, date_logged_out))
        connection.commit()
        return jsonify({"success": True, "message": "User log information saved successfully"})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error saving user log information: {e}"}), 500
    finally:
        cursor.close()
        connection.close()

@app.route('/submit_comment_form', methods=['POST'])
def submit_comment_form():
    required_fields = ["comment_type", "pile_id", "user_id", "date_posted", "comment_text"]
    data = {field: request.form.get(field) for field in required_fields}

    if not all(data.values()):
        return jsonify({"success": False, "message": "Required fields are missing!"}), 400

    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    cursor = connection.cursor()

    try:
        query = """
        INSERT INTO `Comments` (`Comment Type`, `Related Comment ID`, `Pile ID`, `User ID`, `Usage ID`, 
                               `Date Posted`, `Comment Text`, `Comment Date`, `Commented By`, `Status`)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(query, (
            data["comment_type"], request.form.get("related_comment_id"), data["pile_id"], data["user_id"],
            request.form.get("usage_id"), data["date_posted"], data["comment_text"], request.form.get("comment_date"),
            request.form.get("commented_by"), request.form.get("status")
        ))
        connection.commit()
        return jsonify({"success": True, "message": "Comment information saved successfully"})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error saving comment information: {e}"}), 500
    finally:
        cursor.close()
        connection.close()

@app.route('/submit_task_assignment', methods=['POST'])
def submit_task_assignment():
    area_id = request.form.get('area_id')
    user_id = request.form.get('user_id')
    table_id = request.form.get('selectedHotspots')  # Comma-separated Table IDs
    task_date = request.form.get('task_date')
    allotted_date = request.form.get('allotted_date')
    allotted_by = request.form.get('allotted_by')
    date_completed = request.form.get('date_completed')

    if not table_id:
        return jsonify({"success": False, "message": "No tables selected!"}), 400

    table_id_list = [table.strip() for table in table_id.split(",") if table.strip()]
    
    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    cursor = connection.cursor()

    try:
        for table in table_id_list:
            for pile_no in range(1, 5):  # Create 4 rows per Table ID with Pile No 1-4
                # Fetch the last Assessment ID
                cursor.execute("SELECT `Assessment ID` FROM `assessment` ORDER BY `Assessment ID` DESC LIMIT 1")
                result = cursor.fetchone()
                next_number = int(result[0][2:]) + 1 if result else 1
                new_assessment_id = f"AS{next_number:05d}"

                query = """
                INSERT INTO `assessment` 
                    (`Assessment ID`, `Area ID`, `User ID`, `Table ID`, `Pile No`, `Task Date`, 
                     `Allotted Date`, `Allotted By`, `Date Completed`, `Assessment Status`, `Assessment case`)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(query, (
                    new_assessment_id, area_id, user_id, table, pile_no,
                    task_date, allotted_date, allotted_by, date_completed,
                    "In Progress", "Not Assessed"
                ))

        connection.commit()
        return jsonify({"success": True, "message": "Task assignment saved successfully"})

    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error saving task assignment: {e}"}), 500
    finally:
        cursor.close()
        connection.close()

@app.route('/submit_remedy_form', methods=['POST'])
def submit_remedy_form():
    # Get form data
    area_id = request.form.get('area_id')
    user_id = request.form.get('user_id')
    table_ids = request.form.get('selectedHotspots')  # Comma-separated Table IDs
    task_date = request.form.get('task_date')
    assessed_case = request.form.get('assessed_case')
    allotted_date = request.form.get('allotted_date')
    allotted_by = request.form.get('allotted_by')
    date_completed = request.form.get('date_completed')
    remedy_status = request.form.get('remedy_status') or "In Progress"  # Default to "In Progress"
    remedy_text = request.form.get('remedy_text')

    if not table_ids:
        return jsonify({"success": False, "message": "No tables selected!"}), 400

    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        table_id_list = [table.strip() for table in table_ids.split(",") if table.strip()]

        for table_id in table_id_list:
            for pile_no in range(1, 5):  # Create 4 rows per Table ID with Pile No 1-4
                # Fetch the latest Remedy ID
                cursor.execute("SELECT `Remedy ID` FROM `Remedy` ORDER BY `Remedy ID` DESC LIMIT 1")
                result = cursor.fetchone()

                next_number = int(result[0][2:]) + 1 if result else 1
                new_remedy_id = f"RM{next_number:05d}"

                # Insert each row with increasing Pile No
                query = """
                INSERT INTO `Remedy` (`Remedy ID`, `Area ID`, `User ID`, `Table ID`, `Pile No`, `Task Date`, `Assessed Case`, 
                                      `Allotted Date`, `Allotted By`, `Date Completed`, `Remedy Status`, `Remedy Text`)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(query, (new_remedy_id, area_id, user_id, table_id, pile_no, task_date,  "Not Assessed",
                                       allotted_date, allotted_by, date_completed, "In Progress", remedy_text))

        connection.commit()
        return jsonify({"success": True, "message": "Remedy form submitted successfully"})

    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error saving remedy form: {e}"}), 500
    finally:
        cursor.close()
        connection.close()

@app.route('/update_assessment', methods=['POST'])
def update_assessment_pics():
    try:
        user_id = request.form.get('user_id')
        task_date = request.form.get('task_date')
        table_ids = request.form.get('table_id')
        assessment_status = request.form.get('assessment_status')
        assessment_case = request.form.get('assessment_case')
        date_completed = request.form.get('date_completed')

        if not table_ids or "-" not in table_ids:
            return jsonify({"success": False, "message": "Invalid Table ID format"}), 400

        table_id = table_ids.split("-")[0]
        pile_no = table_ids.split("-")[1][-1]

        if not (user_id and task_date and table_id):
            return jsonify({"success": False, "message": "Missing required fields"}), 400

        connection = get_db_connection()
        cursor = connection.cursor(buffered=True)

        # 🔹 Fetch the Assessment ID and existing picture names
        cursor.execute("""
            SELECT `Assessment ID`, `Picture1 Name`, `Picture2 Name`, `Picture3 Name`, `Picture4 Name`
            FROM assessment 
            WHERE `User ID` = %s AND `Task Date` = %s AND `Table ID` = %s AND `Pile No` = %s
        """, (user_id, task_date, table_id, pile_no))

        result = cursor.fetchone()
        if not result:
            cursor.close()
            connection.close()
            return jsonify({"success": False, "message": "Assessment not found"}), 404

        assessment_id, pic1_old, pic2_old, pic3_old, pic4_old = result
        cursor.close()

        assessment_folder = os.path.join(app.config["UPLOAD_FOLDER"], f"{assessment_id}")
        os.makedirs(assessment_folder, exist_ok=True)
        print(f"Saving images to folder: {assessment_folder}")

        # 🔹 Fetch Pile IDs (optional for validation)
        cursor = connection.cursor(buffered=True)
        cursor.execute("""
            SELECT `Pile ID` FROM piles 
            WHERE `Table ID` = %s 
            ORDER BY `Pile No` ASC
        """, (table_id,))
        pile_ids = [row[0] for row in cursor.fetchall()]
        cursor.close()

        if len(pile_ids) < 4:
            connection.close()
            return jsonify({"success": False, "message": "Not enough Pile IDs found"}), 400

        # 🔹 Save new images if uploaded (overwrite logic)
        image_paths = [pic1_old, pic2_old, pic3_old, pic4_old]
        for i in range(4):
            image_field = f'image{i+1}'
            new_image = request.files.get(image_field)

            if new_image:
                image_filename = f"{assessment_id}_{table_id}_Pile{pile_no}_side{i+1}.jpg"
                image_path = os.path.join(assessment_folder, image_filename)
                new_image.save(image_path)
                image_paths[i] = image_filename
                print(f"Saved {image_field} to {image_path}")

        # 🔹 Update the database with new picture names and metadata
        cursor = connection.cursor()
        update_query = """
            UPDATE assessment 
            SET `Assessment Status` = %s, `Assessment Case` = %s, `Date Completed` = %s,
                `Picture1 Name` = %s, `Picture2 Name` = %s, 
                `Picture3 Name` = %s, `Picture4 Name` = %s,
                `Picture Location` = %s
            WHERE `Assessment ID` = %s
        """
        cursor.execute(update_query, (
            assessment_status, assessment_case, date_completed,
            image_paths[0], image_paths[1], image_paths[2], image_paths[3],
            assessment_folder,
            assessment_id
        ))

        connection.commit()
        cursor.close()
        connection.close()

        return jsonify({
            "success": True,
            "message": "Assessment updated successfully",
            "folder": assessment_folder
        })

    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Database error: {str(e)}"}), 500

    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/update_remedy_pics', methods=['POST'])
def update_remedy_pics():
    try:
        # Get form data
        user_id = request.form.get('user_id')
        task_date = request.form.get('task_date')
        table_ids = request.form.get('table_id')
        assessed_case = request.form.get('assessed_case')  
        remedy_status = request.form.get('remedy_status')
        date_completed = request.form.get('date_completed')

        if not table_ids or "-" not in table_ids:
            return jsonify({"success": False, "message": "Invalid Table ID format"}), 400

        table_id = table_ids.split("-")[0]
        pile_no = table_ids.split("-")[1][-1]

        if not (user_id and task_date and table_id and pile_no):
            return jsonify({"success": False, "message": "Missing required fields"}), 400

        # Get the corresponding Remedy ID
        connection = get_db_connection()
        cursor = connection.cursor(buffered=True)

        cursor.execute("""
            SELECT `Remedy ID`, `Picture1 Name`, `Picture2 Name`, `Picture3 Name`, `Picture4 Name`
            FROM remedy 
            WHERE `User ID` = %s AND `Task Date` = %s AND `Table ID` = %s AND `Pile No` = %s
        """, (user_id, task_date, table_id, pile_no))

        result = cursor.fetchone()
        cursor.close()

        if not result:
            connection.close()
            return jsonify({"success": False, "message": "Remedy not found"}), 404

        remedy_id = result[0]

        # Create a unique folder for the remedy
        remedy_folder = os.path.join(app.config["UPLOAD_IMG"], f"{remedy_id}")
        os.makedirs(remedy_folder, exist_ok=True)

        image_paths = []
        for i in range(4):
            image = request.files.get(f'image{i+1}')
            if image:
                image_filename = f"{remedy_id}_{table_id}_Pile{pile_no}_side{i+1}.jpg"
                image_path = os.path.join(remedy_folder, image_filename)
                image.save(image_path)
                image_paths.append(image_filename)
            else:
                image_paths.append(None)

        # Update database
        cursor = connection.cursor()
        query = """
            UPDATE remedy 
            SET `Remedy Status` = %s, `Assessed Case` = %s, `Date Completed` = %s,
                `Picture1 Name` = %s, `Picture2 Name` = %s, 
                `Picture3 Name` = %s, `Picture4 Name` = %s,
                `Picture Location` = %s
            WHERE `Remedy ID` = %s
        """

        cursor.execute(query, (
            remedy_status, assessed_case, date_completed,
            image_paths[0], image_paths[1], image_paths[2], image_paths[3],
            remedy_folder, remedy_id
        ))

        connection.commit()
        cursor.close()
        connection.close()

        return jsonify({"success": True, "message": "Remedy updated successfully", "folder": remedy_folder})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


   
@app.route('/get_submitted_hotspots', methods=['GET'])
def get_submitted_hotspots():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("SELECT `Table ID` FROM `assessment`")  
        rows = cursor.fetchall()

        # Flatten and split the stored values
        submitted_hotspots = []
        for row in rows:
            if row['Table ID']:
                submitted_hotspots.extend(row['Table ID'].split(","))  # Split on commas

        return jsonify({"submitted_hotspots": submitted_hotspots})  
    finally:
        cursor.close()
        connection.close()


@app.route('/get_final_submitted_hotspots', methods=['GET'])
def get_final_submitted_hotspots():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("SELECT `Table ID` FROM `assessment` WHERE `Assessment Status` = 'OE Approved'")  
        rows = cursor.fetchall()

        # Flatten and split the stored values
        final_submitted_hotspots = []
        for row in rows:
            if row['Table ID']:
                final_submitted_hotspots.extend(row['Table ID'].split(","))  # Split on commas

        return jsonify({"final_submitted_hotspots": final_submitted_hotspots})  
    finally:
        cursor.close()
        connection.close()

@app.route('/get_submitted_hotspots_remedy', methods=['GET'])
def get_submitted_hotspots_remedy():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("SELECT `Table ID` FROM `remedy`")  
        rows = cursor.fetchall()

        # Flatten and split the stored values
        submitted_hotspots = []
        for row in rows:
            if row['Table ID']:
                submitted_hotspots.extend(row['Table ID'].split(","))  # Split on commas

        return jsonify({"submitted_hotspots": submitted_hotspots})  
    finally:
        cursor.close()
        connection.close()

@app.route('/submit_row_form', methods=['POST'])
def submit_row_form():
    # Retrieve form data
    row_name = request.form.get('row_name')
    area_id = request.form.get('area_id')
    location = request.form.get('location')
    gps = request.form.get('gps')

    # Validate required fields
    if not row_name or not area_id or not location or not gps:
        return jsonify({"success": False, "message": "All fields are required."})

    # Establish DB connection
    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        # Fetch the current maximum Row ID
        cursor.execute("SELECT `Row ID` FROM `rows` ORDER BY `Row ID` DESC LIMIT 1")
        result = cursor.fetchone()

        # Determine the next Row ID
        if result and result[0]:
            last_row_id = result[0]
            next_number = int(last_row_id[1:]) + 1
        else:
            next_number = 1

        # Format the new Row ID as 'R001', 'R002', etc.
        new_row_id = f"R{next_number:03d}"

        # Insert data into the rows table with the generated Row ID
        query = """
        INSERT INTO `rows` (`Row ID`, `Row Name`, `Area ID`, `Location`, `GPS`)
        VALUES (%s, %s, %s, %s, %s)
        """
        cursor.execute(query, (new_row_id, row_name, area_id, location, gps))
        connection.commit()

        return jsonify({"success": True, "message": f"Row information saved successfully with ID {new_row_id}"})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error saving row information: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route('/submit_pile_form', methods=['POST'])
def submit_pile_form():
    # Get form data
    pile_ids = request.form.get('table_id')  # This contains multiple pile IDs as a string (e.g., "P1, P2, P3")
    area_id = request.form.get('area_id')
    location_description = request.form.get('location_description')
    gps_location = request.form.get('gps_location')

    # Establish DB connection
    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        if not pile_ids:
            return jsonify({"success": False, "message": "No pile selected"})
        pile_list = pile_ids.split(", ")  
        for pile_id in pile_list:
            # Fetch the current maximum Pile ID modifiy the
            cursor.execute("SELECT `Pile ID` FROM `Piles` ORDER BY `Pile ID` DESC LIMIT 1")
            result = cursor.fetchone()

            # Generate the next Pile ID
            if result and result[0]:
                last_pile_id = result[0]
                next_number = int(last_pile_id[1:]) + 1  
            else:
                next_number = 1  

            new_pile_id = f"P{next_number:03d}"

            # Insert each pile as a new row
            query = """
            INSERT INTO `Piles` (`Pile ID`, `Table ID`,  `Area ID`, `Location Description`, `GPS Location`)
            VALUES (%s, %s, %s, %s, %s)
            """
            cursor.execute(query, (new_pile_id, pile_id,  area_id, location_description, gps_location))

        connection.commit()

        return jsonify({"success": True, "message": "Pile information saved successfully"})

    except mysql.connector.Error as e:
        print(f"Error: {e}")
        return jsonify({"success": False, "message": f"Error saving pile information: {e}"})

    finally:
        cursor.close()
        connection.close()

@app.route('/submit_table_form', methods=['POST'])
def submit_table_form():
    # Get form data
    table_ids = request.form.get('table_id')  # This contains multiple table IDs as a string (e.g., "T1, T2, T3")
    area_id = request.form.get('area_id')
    location_description = request.form.get('location')
    gps_location = request.form.get('gps_location')

    # Establish DB connection
    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        if not table_ids:
            return jsonify({"success": False, "message": "No table IDs provided"})

        # Split the comma-separated table IDs into a list
        table_list = table_ids.split(", ")  # Convert the string of table_ids to a list

        # Insert each table as a new row
        for table_id in table_list:
            # Insert into the `tables` table
            query = """
            INSERT INTO `tables` (`Table ID`, `Area ID`, `Location`, `GPS`)
            VALUES (%s, %s, %s, %s)
            """
            cursor.execute(query, (table_id, area_id, location_description, gps_location))

        connection.commit()

        return jsonify({"success": True, "message": "Table information saved successfully"})

    except mysql.connector.Error as e:
        print(f"Error: {e}")
        return jsonify({"success": False, "message": f"Error saving table information: {e}"})

    finally:
        cursor.close()
        connection.close()

@app.route('/submit_inventory_details', methods=['POST'])
def submit_inventory_details():
    # Get form data
    item_type = request.form.get('item_type')
    item_uom = request.form.get('item_uom')
    item_desc = request.form.get('item_desc')
    item_avl_qty = request.form.get('item_avl_qty')
    item_ror = request.form.get('item_ror') or None
    item_value = request.form.get('item_value') or None
    item_rate = request.form.get('item_rate') or None

    # Establish DB connection
    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        # Generate the next Item ID
        cursor.execute("SELECT `Item ID` FROM `Inventory` ORDER BY `Item ID` DESC LIMIT 1")
        result = cursor.fetchone()
        
        if result and result[0]:
            last_item_id = result[0]
            next_number = int(last_item_id[1:]) + 1
        else:
            next_number = 1
        
        # Format the new Item ID as 'I001', 'I002', etc.
        new_item_id = f"I{next_number:04d}"

        # Insert data into Inventory table
        query = """
        INSERT INTO `Inventory` (`Item ID`, `Item Type`, `Item UOM`, `Item Desc`, 
                                 `Item Avl Qty`, `Item ROR`, `Item Value`, `Item Rate`)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(query, (new_item_id, item_type, item_uom, item_desc, item_avl_qty, item_ror, item_value, item_rate))
        connection.commit()

     
        return jsonify({"success": True, "message": f"Item details saved successfully "})

    except mysql.connector.Error as e:
        print(f"Error: {e}")
        return jsonify({"success": False, "message": f"Error saving item details: {e}"})

    finally:
        cursor.close()
        connection.close()

#invtrans
@app.route('/submit_item_transaction_form', methods=['POST'])
def submit_item_transaction_form():
    # Get form data
    item_type = request.form.get('item_type')
    trans_qty = request.form.get('trans_qty')
    trans_type = request.form.get('trans_type')
    trans_date = request.form.get('trans_date')
    user_id = request.form.get('user_id')
    usage = request.form.get('usage')

    # Establish DB connection
    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        # Get the last Item ID
        cursor.execute("SELECT `Item ID` FROM `Invtrans` ORDER BY `Item ID` DESC LIMIT 1")
        result = cursor.fetchone()

        if result and result[0]:  # Ensure result exists and is not None
            last_item_id = result[0]

            if last_item_id.startswith("IT"):  # Check format
                try:
                    next_number = int(last_item_id[2:]) + 1  # Extract number part safely
                except ValueError:
                    next_number = 1  # Reset if format is incorrect
            else:
                next_number = 1  # Reset if format is not as expected
        else:
            next_number = 1  # If no previous records exist, start with IT001

        # Generate new Item ID
        new_item_id = f"IT{next_number:04d}"  # Format as IT001, IT002, etc.

        # Insert data into Item Transaction table
        query = """
        INSERT INTO `invtrans` (`Item ID`, `Item Type`, `Trans Qty`, `Trans Type`, 
                                `Trans Date`, `User ID`, `Usage`)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(query, (new_item_id, item_type, trans_qty, trans_type, trans_date, user_id, usage))
        connection.commit()

     
        return jsonify({"success": True, "message": "Item transaction details saved successfully"})

    except mysql.connector.Error as e:
        print(f"Error: {e}")
        return jsonify({"success": False, "message": f"Error saving item transaction details: {e}"})

    finally:
        cursor.close()
        connection.close()


###########################################################################

@app.route('/generate_report', methods=['GET'])
def generate_report():
    user_id = request.args.get('user_id')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    assessment_status = request.args.get('assessment_status')
    area_id = request.args.get('area')  # ✅ Get the Area ID

    if not from_date or not to_date:
        return jsonify({"error": "Missing required parameters"}), 400

    connection = get_db_connection()
    if not connection:
        return jsonify({"error": "Database connection failed"}), 500

    cursor = connection.cursor(dictionary=True)

    try:
        query = """
            SELECT 
                a.`Assessment ID`, 
                u.`User Name`, 
                a.`Table ID`, 
                a.`Pile No`, 
                DATE_FORMAT(a.`Task Date`, '%d %b %Y') AS `Task Date`,
                a.`Assessment Status`, 
                a.`Assessment Case`
            FROM assessment a
            JOIN users u ON a.`User ID` = u.`User ID`
            WHERE DATE(a.`Task Date`) BETWEEN %s AND %s
        """
        params = [from_date, to_date]

        if user_id and user_id.lower() != "all":
            query += " AND a.`User ID` = %s"
            params.append(user_id)

        if assessment_status and assessment_status.lower() != "all status":
            query += " AND a.`Assessment Status` = %s"
            params.append(assessment_status)

        if area_id and area_id.lower() != "all":
            query += " AND a.`Area ID` = %s"
            params.append(area_id)

        cursor.execute(query, params)
        result = cursor.fetchall()

        return jsonify(result)

    except mysql.connector.Error as e:
        return jsonify({"error": f"Database error: {e}"}), 500

    finally:
        cursor.close()
        connection.close()


@app.route('/generate_remedy_report', methods=['GET'])
def generate_remedy_report():
    user_id = request.args.get('user_id')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    remedy_status = request.args.get('remedy_status')
    area_id = request.args.get('area')  # Get Area ID from request

    if not from_date or not to_date:
        return jsonify({"error": "Missing required parameters"}), 400

    connection = get_db_connection()
    if not connection:
        return jsonify({"error": "Database connection failed"}), 500

    cursor = connection.cursor(dictionary=True)

    try:
        query = """
            SELECT a.`Remedy ID`, u.`User Name`, a.`Table ID`, a.`Pile No`, 
       DATE_FORMAT(a.`Task Date`, '%d %b %Y') AS `Task Date`,
       a.`Remedy Status`, a.`Assessed Case`, a.`Area ID`,
       a.`Picture1 Name`, a.`Picture2 Name`, a.`Picture3 Name`, a.`Picture4 Name`
FROM remedy a
JOIN users u ON a.`User ID` = u.`User ID`
WHERE DATE(a.`Task Date`) BETWEEN %s AND %s
        """
        params = [from_date, to_date]

        # Filter by User ID (if not "all")
        if user_id and user_id.lower() != "all":
            query += " AND a.`User ID` = %s"
            params.append(user_id)

        # Filter by Remedy Status (if not "all status")
        if remedy_status and remedy_status.lower() != "all status":
            query += " AND a.`Remedy Status` = %s"
            params.append(remedy_status)

        # Filter by Area ID (if selected)
        if area_id and area_id.lower() != "all":
            query += " AND a.`Area ID` = %s"
            params.append(area_id)

        cursor.execute(query, params)
        result = cursor.fetchall()

        return jsonify(result)

    except mysql.connector.Error as e:
        return jsonify({"error": f"Database error: {e}"}), 500

    finally:
        cursor.close()
        connection.close()


@app.route('/save_pdf', methods=['POST'])
def save_pdf():
    pdf_file = request.files['pdf']

    # Get Desktop Path
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

    # Create "Assessment Reports" Folder if it doesn't exist
    reports_folder = os.path.join(desktop_path, "Assessment Reports")
    os.makedirs(reports_folder, exist_ok=True)

    # Extract base name and extension
    original_filename = pdf_file.filename
    base_name, extension = os.path.splitext(original_filename)

    # Prepare file path and check for duplicates
    pdf_path = os.path.join(reports_folder, original_filename)
    count = 1
    while os.path.exists(pdf_path):
        pdf_path = os.path.join(reports_folder, f"{base_name}_{count}{extension}")
        count += 1

    # Save the file with a unique name
    pdf_file.save(pdf_path)

    return jsonify({"message": f"PDF saved successfully as {os.path.basename(pdf_path)}!"})
    
slno_file = "last_slno.txt"

# Helper to read last slno
def get_last_slno():
    if not os.path.exists(slno_file):
        return 0
    with open(slno_file, "r") as f:
        return int(f.read().strip())

# Helper to update last slno
def update_last_slno(new_slno):
    with open(slno_file, "w") as f:
        f.write(str(new_slno))

@app.route('/get_last_slno', methods=['GET'])
def get_slno():
    return jsonify({"last_slno": get_last_slno()})

@app.route('/savecase_pdf', methods=['POST'])
def savecase_pdf():
    pdf_file = request.files['pdf']
    new_last_slno = int(request.form.get("new_last_slno", "0"))

    # Save to casereports folder on Desktop
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    reports_folder = os.path.join(desktop_path, "Assessment Reports", "Case Reports")
    os.makedirs(reports_folder, exist_ok=True)

    original_filename = pdf_file.filename
    base_name, extension = os.path.splitext(original_filename)
    pdf_path = os.path.join(reports_folder, original_filename)
    count = 1
    while os.path.exists(pdf_path):
        pdf_path = os.path.join(reports_folder, f"{base_name}_{count}{extension}")
        count += 1

    pdf_file.save(pdf_path)

    # 🔁 Update the Sl. No.
    update_last_slno(new_last_slno)

    return jsonify({"message": f"PDF saved as {os.path.basename(pdf_path)}!"})

@app.route('/save_remedypdf', methods=['POST'])
def save_remedypdf():
    pdf_file = request.files['pdf']

    # Get Desktop Path
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

    # Create "Remedy Reports" Folder if it doesn't exist
    reports_folder = os.path.join(desktop_path, "Remedy Reports")
    os.makedirs(reports_folder, exist_ok=True)

    # Extract base name and extension
    original_filename = pdf_file.filename
    base_name, extension = os.path.splitext(original_filename)

    # Prepare file path and check for duplicates
    pdf_path = os.path.join(reports_folder, original_filename)
    count = 1
    while os.path.exists(pdf_path):
        pdf_path = os.path.join(reports_folder, f"{base_name}_{count}{extension}")
        count += 1

    # Save the file with a unique name
    pdf_file.save(pdf_path)

    return jsonify({"message": f"PDF saved successfully as {os.path.basename(pdf_path)}!"})

@app.route('/get_user_ids', methods=['GET'])
def get_user_ids():
    if "user_type" not in session:
        return jsonify({"success": False, "message": "User not authenticated"}), 401

    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        if session["user_type"] == "Admin":
            # Admin can see all users
            cursor.execute("SELECT `User ID`, `User Name` FROM `users`")
        else:
            # Normal user can see only their own details
            cursor.execute("SELECT `User ID`, `User Name` FROM `users` WHERE `Email` = %s", (session["email"],))

        users = [{"id": row[0], "username": row[1]} for row in cursor.fetchall()]
        
        return jsonify({"success": True, "users": users})

    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching User IDs: {e}"})

    finally:
        cursor.close()
        connection.close()


@app.route('/get_site_ids', methods=['GET'])
def get_site_ids():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        cursor.execute("SELECT `Site ID`, `Site Name` FROM `site`")
        sites = [{"id": row[0], "name": row[1]} for row in cursor.fetchall()]
        return jsonify({"success": True, "sites": sites})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching Site IDs and Site Names: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route('/get_area_ids', methods=['GET'])
def get_area_ids():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        # Fetch Area IDs and Locations
        cursor.execute("SELECT `Area ID`, `Location` FROM `areas`")
        areas = [{"id": row[0], "location": row[1]} for row in cursor.fetchall()]
        return jsonify({"success": True, "areas": areas})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching Area IDs and Locations: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route("/get_row_ids", methods=["GET"])
def get_row_ids():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        # Fetch Row IDs and Row Names from the database
        cursor.execute("SELECT `Row ID`, `Row Name` FROM `rows`")
        rows = [{"id": row[0], "name": row[1]} for row in cursor.fetchall()]
        return jsonify({"success": True, "rows": rows})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching rows: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route("/get_table_ids", methods=["GET"])
def get_table_ids():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        # Fetch Table IDs and Locations from the database
        cursor.execute("SELECT `Table ID`, `Location` FROM `tables`")
        tables = [{"id": table[0], "location": table[1]} for table in cursor.fetchall()]
        return jsonify({"success": True, "tables": tables})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching tables: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route('/get_customer_ids', methods=['GET'])
def get_customer_ids():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        # Fetch Customer ID and Customer Name from the database
        cursor.execute("SELECT `Cust ID`, `Customer Name` FROM `customer`")
        customers = [{"id": row[0], "name": row[1]} for row in cursor.fetchall()]
        
        # Return the list of customers as a JSON response
        return jsonify({"success": True, "customers": customers})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching Customer IDs and Names: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route('/get_item_names', methods=['GET'])
def get_item_names():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        query = "SELECT `Item ID`, `Item Type` FROM `inventory`"
        cursor.execute(query)
        items = cursor.fetchall()

        item_list = [{"item_id": item[0], "item_name": item[1]} for item in items]

        return jsonify({"success": True, "items": item_list})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching item names: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route('/get_itemtrns_names', methods=['GET'])
def get_itemtrns_names():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        query = "SELECT `Item ID`, `Item Type` FROM `invtrans`"
        cursor.execute(query)
        items = cursor.fetchall()

        item_list = [{"item_id": item[0], "item_name": item[1]} for item in items]

        return jsonify({"success": True, "items": item_list})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching item names: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route('/search', methods=['GET'])
def search():
    query = request.args.get('query', '')  # Get the search query from the request
    results = []

    try:
        # Connect to the database using get_db_connection
        connection = get_db_connection()
        cursor = connection.cursor(dictionary=True)

        # SQL Query to search for assessment names
        sql_query = """
        SELECT `assessment id`
        FROM assessment
        WHERE `assessment id` LIKE %s
        LIMIT 10
        """ 
        cursor.execute(sql_query, (f"%{query}%",))

        # Fetch all results
        results = cursor.fetchall()

    except mysql.connector.Error as err:
        print(f"Error: {err}")
        return jsonify({"error": str(err)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

    return jsonify({"results": results})


@app.route('/search_by_date', methods=['GET'])
def search_by_date():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    user_id = request.args.get('user_id')
    task_date = request.args.get('date')

    if not user_id or not task_date:
        return jsonify({"error": "User ID and Date parameters are required"}), 400

    try:
        # Fetch Table ID, Assessment Status, Assessment Case, and Pile No separately
        cursor.execute("""
            SELECT `Table ID`, `Assessment Status`, `Assessment Case`, `Pile No`
            FROM assessment
            WHERE `User ID` = %s 
            AND DATE(`Task Date`) = %s  
            
            ORDER BY `Table ID`, `Pile No`
        """, (user_id, task_date))

        result = cursor.fetchall()

        formatted_data = []
        for row in result:
            table_id = row["Table ID"]
            pile_no = row["Pile No"]

            formatted_data.append({
                "Assessment Case": row["Assessment Case"],
                "Assessment Status": row["Assessment Status"],
                "Pile No": pile_no,
                "Table ID": f"{table_id}-Pile{pile_no}"  # Format Table ID with Pile No
            })

        return jsonify({"data": formatted_data}) if formatted_data else jsonify({"data": []})

    except mysql.connector.Error as e:
        return jsonify({"error": f"Database error: {e}"}), 500

    finally:
        cursor.close()
        connection.close()

@app.route('/search_by_remedydate', methods=['GET'])
def search_by_remedydate():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    
    user_id = request.args.get('user_id')
    task_date = request.args.get('date')

    if not user_id or not task_date:
        return jsonify({"error": "User ID and Date parameters are required"}), 400

    try:
        # Fetch Table ID, Pile No, Remedy Status, and Assessed Case
        cursor.execute("""
            SELECT `Table ID`, `Pile No`, `Remedy Status`, `Assessed Case`
            FROM remedy
            WHERE `User ID` = %s 
            AND DATE(`Task Date`) = %s  
            AND (`Remedy Status` != 'OE Approved' OR `Remedy Status` IS NULL)
            ORDER BY `Table ID`, `Pile No`
        """, (user_id, task_date))

        result = cursor.fetchall()

        formatted_data = []
        for row in result:
            table_id = row["Table ID"]
            pile_no = row["Pile No"]

            formatted_data.append({
                "Assessed Case": row["Assessed Case"],
                "Remedy Status": row["Remedy Status"],
                "Pile No": pile_no,
                "Table ID": f"{table_id}-Pile{pile_no}"  # Format Table ID with Pile No
            })

        return jsonify({"data": formatted_data}) if formatted_data else jsonify({"data": []})

    except mysql.connector.Error as e:
        return jsonify({"error": f"Database error: {e}"}), 500

    finally:
        cursor.close()
        connection.close()


############################################################################

@app.route('/submit_updateuser_form', methods=['POST'])
def submit_updateuser_form():
    user_id = request.form.get('user_id')  
    user_email = request.form.get('user_email')  
    user_password = request.form.get('user_password')  
    phone_no = request.form.get('phone_no')
    date_removed = request.form.get('date_removed')
    user_type = request.form.get('user_type')  # Capture the user type from the form

    # Connect to the database
    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        update_fields = []
        params = []

        if user_email:
            update_fields.append("`Email` = %s")
            params.append(user_email)

        if user_password:
            update_fields.append("`Password` = %s")
            params.append(user_password)

        if phone_no:
            update_fields.append("`User Phone number` = %s")
            params.append(phone_no)

        if date_removed:
            update_fields.append("`Date removed` = %s")
            params.append(date_removed)

        if user_type:  # Add a condition for user_type
            update_fields.append("`User Type` = %s")
            params.append(user_type)

        if not update_fields:
            return jsonify({"success": False, "message": "No fields to update provided."})

        params.append(user_id)
        query = f"""
            UPDATE `users`
            SET {', '.join(update_fields)}
            WHERE `User ID` = %s
        """
        cursor.execute(query, tuple(params))
        connection.commit()

        return jsonify({"success": True, "message": "User Details updated successfully."})

    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error updating user: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route('/delete_user/<user_id>', methods=['DELETE'])
def delete_user(user_id):
    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        # Check if the user exists
        cursor.execute("SELECT `User ID` FROM `users` WHERE `User ID` = %s", (user_id,))
        if not cursor.fetchone():
            return jsonify({"success": False, "message": "User not found."})

        # Delete the user
        cursor.execute("DELETE FROM `users` WHERE `User ID` = %s", (user_id,))
        connection.commit()

        return jsonify({"success": True, "message": "User deleted successfully."})

    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error deleting user: {e}"})
    finally:
        cursor.close()
        connection.close()

# Update site details
@app.route("/submit_site_update", methods=["POST"])
def submit_site_update():
    try:
        data = request.form
        site_id = data.get("site_id")
        site_location = data.get("location")
        site_owner_name = data.get("site_owner_name")

        if not site_id or not site_location or not site_owner_name:
            return jsonify({"success": False, "message": "All fields are required."})

        connection = get_db_connection()
        cursor = connection.cursor()

        # Enclose column names with spaces in backticks
        query = """
            UPDATE `site` 
            SET `Site Location` = %s, `Site Owner Name` = %s 
            WHERE `Site ID` = %s
        """
        cursor.execute(query, (site_location, site_owner_name, site_id))
        connection.commit()

        if cursor.rowcount > 0:
            return jsonify({"success": True, "message": "Site updated successfully."})
        else:
            return jsonify({"success": False, "message": "No changes made or site not found."})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})
    finally:
        cursor.close()
        connection.close()

# Delete a site
@app.route("/delete_site/<site_id>", methods=["DELETE"])
def delete_site(site_id):
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Use backticks for column names
        query = "DELETE FROM `site` WHERE `Site ID` = %s"
        cursor.execute(query, (site_id,))
        connection.commit()

        if cursor.rowcount > 0:
            return jsonify({"success": True, "message": "Site deleted successfully."})
        else:
            return jsonify({"success": False, "message": "Site not found."})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})
    finally:
        cursor.close()
        connection.close()

# Route to handle customer update
@app.route('/submit_customer_update', methods=['POST'])
def submit_customer_update():
    try:
        # Get form data
        customer_id = request.form.get('customer_name')
        address = request.form.get('address')
        phone_no = request.form.get('phone_no')

        if not customer_id:
            return jsonify({"success": False, "message": "Customer ID is required."})

        # Initialize list to store the updates
        updates = []
        values = []

        # Only add fields if they are provided
        if address:
            updates.append("`Customer Address` = %s")
            values.append(address)
        if phone_no:
            updates.append("`Phone No` = %s")
            values.append(phone_no)

        # If no updates are provided, return an error
        if not updates:
            return jsonify({"success": False, "message": "At least one field should be provided to update."})

        # Add the customer_id to the end of values
        values.append(customer_id)

        # Create the SQL query dynamically based on provided fields
        query = f"""
            UPDATE `customer`
            SET {', '.join(updates)}
            WHERE `Cust ID` = %s
        """

        # Get database connection
        connection = get_db_connection()
        cursor = connection.cursor()

        # Execute the update query
        cursor.execute(query, tuple(values))
        connection.commit()

        # Check if the update was successful
        if cursor.rowcount > 0:
            return jsonify({"success": True, "message": "Customer updated successfully."})
        else:
            return jsonify({"success": False, "message": "No changes made or customer not found."})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})
    finally:
        # Close the cursor and connection
        cursor.close()
        connection.close()

# Route to handle customer deletion
@app.route("/delete_customer/<customer_id>", methods=["DELETE"])
def delete_customer(customer_id):
    try:
        connection = get_db_connection()
        cursor = connection.cursor()

        # Delete query with backticks for columns
        query = "DELETE FROM `customer` WHERE `Cust ID` = %s"
        cursor.execute(query, (customer_id,))
        connection.commit()

        if cursor.rowcount > 0:
            return jsonify({"success": True, "message": "Customer deleted successfully."})
        else:
            return jsonify({"success": False, "message": "Customer not found."})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})
    finally:
        cursor.close()
        connection.close()

@app.route('/get_customer_details/<customer_id>', methods=['GET'])
def get_customer_details(customer_id):
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
       
        query = """
        SELECT `Customer Address`, `Phone No`
        FROM `Customer`
        WHERE `Cust ID` = %s
        """
        cursor.execute(query, (customer_id,))
        customer = cursor.fetchone()

        if customer:
           
            return jsonify({"success": True, "customer": {
                "address": customer[0],
                "phone_no": customer[1]
            }})
        else:
            return jsonify({"success": False, "message": "Customer not found."})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching customer details: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route('/get_inventory_details/<item_id>', methods=['GET'])
def get_inventory_details(item_id):
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        query = """
        SELECT `Item Type`, `Item UOM`, `Item Avl Qty`, `Item ROR`, `Item Value`, `Item Rate`
        FROM `inventory`
        WHERE `Item ID` = %s
        """
        cursor.execute(query, (item_id,))
        item = cursor.fetchone()

        if item:
            return jsonify({
                "success": True,
                "inventory": {
                    "item_type": item[0],
                    "item_uom": item[1],
                    "item_avl_qty": item[2],
                    "item_ror": item[3],
                    "item_value": item[4],
                    "item_rate": item[5]
                }
            })
        else:
            return jsonify({"success": False, "message": "Item not found."})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching item details: {e}"})
    finally:
        cursor.close()
        connection.close()


@app.route('/get_site_details/<site_id>', methods=['GET'])
def get_site_details(site_id):
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        # Query to fetch site details (location, site owner name)
        query = """
        SELECT `Site Location`, `Site Owner Name`
        FROM `Site`
        WHERE `Site ID` = %s
        """
        cursor.execute(query, (site_id,))
        site = cursor.fetchone()

        if site:
            # Return site details as JSON
            return jsonify({"success": True, "site": {
                "location": site[0],
                "owner_name": site[1]
            }})
        else:
            return jsonify({"success": False, "message": "Site not found."})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching site details: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route('/get_user_details/<user_id>', methods=['GET'])
def get_user_details(user_id):
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        # Query to fetch user details (email, password, phone number, etc.)
        query = """
        SELECT `Email`, `Password`, `User Phone number`
        FROM `users`
        WHERE `User ID` = %s
        """
        cursor.execute(query, (user_id,))
        user = cursor.fetchone()

        if user:
            # Return user details as JSON
            return jsonify({
                "success": True,
                "user": {
                    "email": user[0],
                    "password": user[1],
                    "phone_no": user[2]
                }
            })
        else:
            return jsonify({"success": False, "message": "User not found."})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching user details: {e}"})
    finally:
        cursor.close()
        connection.close()
@app.route("/update_inventory_details", methods=["POST"])
def update_inventory_details():
    try:
        data = request.form
        item_id = data.get("item_type")  # Should be item_id, not item_type
        if not item_id:
            return jsonify({"success": False, "message": "Item ID is required for updating."})

        column_mapping = {
            "item_uom": "Item UOM",
            "item_desc": "Item Desc",
            "item_avl_qty": "Item Avl Qty",
            "item_ror": "Item ROR",
            "item_value": "Item Value",
            "item_rate": "Item Rate"
        }

        updates = []
        values = []

        for form_field, db_column in column_mapping.items():
            value = data.get(form_field)
            if value:
                updates.append(f"`{db_column}` = %s")
                values.append(value)

        if not updates:
            return jsonify({"success": True, "message": "No fields provided for update. Item is unchanged."})

        query = f"""
            UPDATE `inventory`
            SET {', '.join(updates)}
            WHERE `Item ID` = %s
        """
        values.append(item_id)  # Use Item ID as the identifier

        connection = get_db_connection()
        cursor = connection.cursor()
        cursor.execute(query, tuple(values))
        connection.commit()

        # Always return a success message
        return jsonify({"success": True, "message": "Inventory updated successfully."})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

    finally:
        if "cursor" in locals():
            cursor.close()
        if "connection" in locals():
            connection.close()



@app.route('/delete_item/<item_id>', methods=['DELETE'])
def delete_item(item_id):
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    try:
        # Execute delete query using Item ID
        cursor.execute("DELETE FROM inventory WHERE `Item ID` = %s", (item_id,))
        connection.commit()

        if cursor.rowcount > 0:
            return jsonify({
                'success': True,
                'message': f'Item with ID "{item_id}" deleted successfully.'
            })
        else:
            return jsonify({
                'success': False,
                'message': f'Item with ID "{item_id}" not found.'
            }), 404
    except Exception as e:
        connection.rollback()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500
    finally:
        cursor.close()
        connection.close()

@app.route("/update_item_transaction_form", methods=["POST"])
def update_item_transaction_form():
    try:
        data = request.form
        item_id = data.get("item_type")  # Should be item_id, not item_name
        if not item_id:
            return jsonify({"success": False, "message": "Item ID is required for updating."})

        column_mapping = {
            "trans_qty": "Trans Qty",
            "trans_type": "Trans Type",
            "trans_date": "Trans Date",
            "user_id": "User ID",
            "usage": "Usage"
        }

        updates = []
        values = []

        for form_field, db_column in column_mapping.items():
            value = data.get(form_field)
            if value:
                updates.append(f"`{db_column}` = %s")
                values.append(value)

        if not updates:
            return jsonify({"success": True, "message": "No fields provided for update. Item is unchanged."})

        query = f"""
            UPDATE `invtrans`
            SET {', '.join(updates)}
            WHERE `Item ID` = %s
        """
        values.append(item_id)

        connection = get_db_connection()
        cursor = connection.cursor()
        cursor.execute(query, tuple(values))
        connection.commit()

        return jsonify({"success": True, "message": "Item updated successfully."})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

    finally:
        if "cursor" in locals():
            cursor.close()
        if "connection" in locals():
            connection.close()
            
#invtrns
@app.route('/delete_itemtrns/<item_id>', methods=['DELETE'])
def delete_itemtrns(item_id):
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    try:
        cursor.execute("DELETE FROM invtrans WHERE `Item ID` = %s", (item_id,))
        connection.commit()

        return jsonify({
            'success': True,
            'message': f'Item with ID "{item_id}" deleted successfully.'
        })
    except Exception as e:
        connection.rollback()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500
    finally:
        cursor.close()
        connection.close()

@app.route("/get_item_details/<item_id>", methods=["GET"])
def get_item_details(item_id):
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("SELECT * FROM invtrans WHERE `Item ID` = %s", (item_id,))
        item = cursor.fetchone()

        if item:
            return jsonify({"success": True, "item": item})
        else:
            return jsonify({"success": False, "message": "Item not found."})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

    finally:
        cursor.close()
        connection.close()

#############################################################################
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True, threaded=True)
